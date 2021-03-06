# -*- coding: utf-8 -*-
# Copyright (c) 2012-2013 by Pablo Martín <goinnn@gmail.com>
#
# This software is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This software is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with this software.  If not, see <http://www.gnu.org/licenses/>.


import csv
import collections
import sys

PY3 = sys.version_info[0] == 3

try:
    from cStringIO import StringIO
except ImportError:
    from io import StringIO


def openExcelSheet():
    """ Opens a reference to an Excel WorkBook and Worksheet objects """
    import pyExcelerator
    workbook = pyExcelerator.Workbook()
    worksheet = workbook.add_sheet("Sheet 1")
    return workbook, worksheet


def validateOpts(response):
    """ Returns option values specified, or the default if none """
    titlePresent = False
    linesPerFile = -1
    sepChar = ","
    return titlePresent, linesPerFile, sepChar


def writeExcelHeader(worksheet, titleCols):
    """ Write the header line into the worksheet """
    cno = 0
    for titleCol in titleCols:
        worksheet.write(0, cno, titleCol)
        cno = cno + 1


def writeExcelRow(worksheet, lno, columns):
    """ Write a non-header row into the worksheet """
    cno = 0
    for column in columns:
        worksheet.write(lno, cno, column.decode('utf-8'))
        cno = cno + 1


def closeExcelSheet(response, workbook):
    """ Saves the in-memory WorkBook object into the specified file """
    response.content = workbook.get_biff_data()


def convert_to_excel_pyexcelerator(response):
    titlePresent, linesPerFile, sepChar = validateOpts(response)
    workbook, worksheet = openExcelSheet()
    fno = 0
    lno = 0
    titleCols = []
    content = StringIO(response.content)
    reader = csv.reader(content)
    for line in reader:
        if (lno == 0 and titlePresent):
            if (len(titleCols) == 0):
                titleCols = line
            writeExcelHeader(worksheet, titleCols)
        else:
            writeExcelRow(worksheet, lno, line)
        lno = lno + 1
        if (linesPerFile != -1 and lno >= linesPerFile):
            fno = fno + 1
            lno = 0
    closeExcelSheet(response, workbook)


# A reasonable approximation for column width is based off zero character in
# default font.  Without knowing exact font details it's impossible to
# determine exact auto width.
# http://stackoverflow.com/questions/3154270/python-xlwt-adjusting-column-widths?lq=1
def get_xls_col_width(text, style):
    return int((1 + len(text)) * 256)


def write_xlwt_row(ws, lno, cell_text, cell_widths, style=None):
    """Write row of utf-8 encoded data to worksheet, keeping track of maximum
    column width for each cell.
    """
    import xlwt
    if style is None:
        style = xlwt.Style.default_style
    for cno, utf8_text in enumerate(cell_text):
        cell_text = utf8_text
        if not PY3:
            cell_text = cell_text.decode('utf-8')
        ws.write(lno, cno, cell_text, style)
        cell_widths[cno] = max(cell_widths[cno],
                               get_xls_col_width(cell_text, style))


def convert_to_excel_xlwt(response):
    """Replace HttpResponse csv content with excel formatted data using xlwt
    library.
    """
    import xlwt
    # Styles used in the spreadsheet.  Headings are bold.
    header_font = xlwt.Font()
    header_font.bold = True

    header_style = xlwt.XFStyle()
    header_style.font = header_font

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet 1')

    # Cell width information kept for every column, indexed by column number.
    cell_widths = collections.defaultdict(lambda: 0)
    if PY3:
        content = StringIO(response.content.decode('utf-8').replace('\x00', ''))
    else:
        content = StringIO(response.content)
    reader = csv.reader(content)
    for lno, line in enumerate(reader):
        if lno == 0:
            style = header_style
        else:
            style = None

        write_xlwt_row(ws, lno, line, cell_widths, style)

    # Roughly autosize output column widths based on maximum column size.
    for col, width in cell_widths.items():
        ws.col(col).width = width

    response.content = ''
    wb.save(response)


def write_openpyxl_row(ws, lno, cell_text, cell_widths):
    for cno, cell_text in enumerate(cell_text):
        if not PY3:
            cell_text = cell_text.decode('utf-8')
        ws.cell(column=cno, row=lno).value = cell_text
        cell_widths[cno] = max(
            cell_widths[cno],
            len(cell_text))


def convert_to_excel_openpyxl(response):
    from openpyxl import Workbook
    from openpyxl.cell import get_column_letter
    wb = Workbook()
    ws = wb.get_active_sheet()

    cell_widths = collections.defaultdict(lambda: 0)

    if PY3:
        content = StringIO(response.content.decode('utf-8').replace('\x00', ''))
    else:
        content = StringIO(response.content)
    reader = csv.reader(content)
    for lno, line in enumerate(reader):
        write_openpyxl_row(ws, lno, line, cell_widths)

    # Roughly autosize output column widths based on maximum column size
    # and add bold style for the header
    for i, cell_width in cell_widths.items():
        ws.cell(column=i, row=0).style.font.bold = True
        ws.column_dimensions[get_column_letter(i + 1)].width = cell_width

    response.content = ''
    wb.save(response)


def get_excel_support():
    # Autodetect library to use for xls writing.  Default to xlwt.
    from django.conf import settings
    EXCEL_SUPPORT = getattr(settings, "EXCEL_SUPPORT", None)
    if EXCEL_SUPPORT:
        return EXCEL_SUPPORT
    try:
        import xlwt
        return 'xlwt'
    except ImportError:
        try:
            import pyExcelerator
            return 'pyexcelerator'
        except ImportError:
            try:
                import openpyxl
                return 'openpyxl'
            except ImportError:
                pass


def convert_to_excel(response):
    EXCEL_SUPPORT = get_excel_support()
    if EXCEL_SUPPORT == 'xlwt':
        convert_to_excel_xlwt(response)
    elif EXCEL_SUPPORT == 'pyexcelerator':
        convert_to_excel_pyexcelerator(response)
    elif EXCEL_SUPPORT == 'openpyxl':
        convert_to_excel_openpyxl(response)
    else:
        raise RuntimeError("No support for xls generation available")
